import threading
from queue import Queue
from time import sleep

import xlrd
import os
from xlutils.copy import copy
import xlwt
import openpyxl

'''
    comm:    文件操作工具类
    author : chongmengzhao
    time:    2019年10月24日10:55:54
'''


class FileUtil(object):
    '''
        comm:   获取指定字符在字符串中最后一次的索引位置
        param:  chars[字符] string [字符串]
    '''

    @staticmethod
    def find_last(chars, string):
        last_position = -1
        while True:
            position = chars.find(string, last_position + 1)
            if position == -1:
                return last_position
            last_position = position

    '''
        comm:   获取变量类型
        param:  variate
    '''

    @staticmethod
    def typeof(variate):

        if isinstance(variate, int):
            return "int"
        elif isinstance(variate, str):
            return "str"
        elif isinstance(variate, float):
            return "float"
        elif isinstance(variate, list):
            return "list"
        elif isinstance(variate, tuple):
            return "tuple"
        elif isinstance(variate, dict):
            return "dict"
        elif isinstance(variate, set):
            return "set"

        return None


'''
    comm:    excel操作工具类
    author : chongmengzhao
    time:    2019年10月24日10:55:54
'''


class Excel(object):
    ThreadList = []
    Loc = None
    LocNum = 4
    dir = None
    suffixName = None
    prefixName = None
    Sheet = "Sheet1"
    Field = None
    Data = None
    Book = None
    ExcelFile = None
    Code = "utf-8"
    LastRow = -1
    LastColumn = -1
    CurrentRow = -1
    CurrentColumn = -1

    '''
        excel文件初始化
    '''

    def __init__(self, *value):
        self.Loc = threading.Lock()

        if len(value) > 0:
            self.file_attr(value)

        if len(value) >= 2:
            self.Sheet = value[1]
        if len(value) >= 3:
            self.Data = Queue()
            for i in value[2]:
                self.Data.put(i)
        if len(value) >= 4:
            self.Field = value[3]
        if len(value) >= 5:
            self.LocNum = value[4]
        if (len(value)) != 0:
            self.createFile()

    '''
        comm:  设置文件名，文件类型，文件路径 三大属性
        param：value[文件路径]
    '''

    def file_attr(self, value):
        if FilePyDataProc.typeof(value) == "str":
            param = str(value).replace("\\", "/")
        else:
            param = str(value[0]).replace("\\", "/")

        self.dir = param[:FilePyDataProc.find_last(param, "/") + 1]
        self.suffixName = param[FilePyDataProc.find_last(param, "."):]
        self.prefixName = param.replace(self.dir, "").replace(self.suffixName, "")

    '''
        comm:  创建excel文件
        param: va[元组类型(文件路径,Sheet名称)]
    '''

    def createFile(self, *va):
        if len(va) != 0:
            if len(va[0]) != 0:
                self.file_attr(va[0])
                if len(va[0]) == 2:
                    self.Sheet = va[0][1]

        self.excel_login()

        return self

    '''
        comm: 验证具体创建什么类型的excel
    '''

    def excel_login(self):
        if self.suffixName == ".xls":
            print("开始创建03文件")
            self.create03File()
        if self.suffixName == ".xlsx":
            print("开始创建07文件")
            self.create07File()

    '''
        comm: 创建03版excel文件
    '''

    def create03File(self):
        self.Book = xlwt.Workbook(encoding=self.Code)
        self.ExcelFile = self.Book.add_sheet(self.Sheet)
        self.Book.save(self.dir + self.prefixName + self.suffixName)
        print("03文件创建完毕")
        print("开始添加标题行")
        self.addTitle03()
        print("开始在03文件追加数据")
        self.append03()

    '''
        comm: 创建07版excel文件
    '''

    def create07File(self):
        self.Book = openpyxl.Workbook()
        del_sheet = self.Book["Sheet"]
        self.Book.remove(del_sheet)
        self.ExcelFile = self.Book.create_sheet(self.Sheet)
        self.Book.save(self.dir + self.prefixName + self.suffixName)
        print("07文件创建完毕")
        print("开始添加标题行")
        self.addTitle07()
        print("开始在07文件追加数据")
        self.append07()

    '''
        comm:  添加07版文件标题
        param: title[元组类型(字段1，字段2，字段3，……)]
    '''

    def addTitle07(self, row_id=1, column_id=1, title=()):
        if len(title) != 0:
            self.Field = title

        if self.Field is not None:
            print("添加07标题")
            self.Book = openpyxl.load_workbook(self.dir + self.prefixName + self.suffixName)
            self.Sheet = self.Book.active
            self.LastRow = row_id
            self.LastColumn = column_id
            self.CurrentRow = row_id
            self.CurrentColumn = column_id
            for j in range(0, len(self.Field)):
                self.Sheet.cell(self.LastRow, self.LastColumn, self.Field[j])
                self.LastColumn = self.LastColumn + 1
            self.Book.save(self.dir + self.prefixName + self.suffixName)
            self.LastRow = self.LastRow + 1

        else:
            print("无可添加的标题")

    '''
        comm:  07文件追加数据
        param: data[元组类型([DataRow],(字段1，字段2，字段3，……),文件路径,开始行号，开始列号)]
    '''

    def append07(self, *data):
        appendFileDir = self.dir + self.prefixName + self.suffixName

        if len(data) == 1:
            self.Data = data[0][0]

        if len(data) == 4:
            self.LastRow = data[0]
            self.LastColumn = data[1]
            self.CurrentRow = data[0]
            self.CurrentColumn = data[1]
            self.LocNum = data[2]
            self.Data = Queue()
            for j in data[3][0]:
                self.Data.put(j)

        if len(data) == 6:
            self.Data = Queue()
            for j in data[0]:
                self.Data.put(j)
            self.Field = data[1]
            appendFileDir = data[2]
            self.LastRow = data[3]
            self.LastColumn = data[4]
            self.CurrentRow = data[3]
            self.CurrentColumn = data[4]
            self.LocNum = data[5]

        if self.Field is not None and self.Data is not None:
            print("添加07文件的行级数据")
            self.Book = openpyxl.load_workbook(appendFileDir)
            self.Sheet = self.Book.active

            for j in range(self.LocNum):
                excel = threading.Thread(target=self.write07Excel, args=(data,))
                self.ThreadList.append(excel)

            for i in self.ThreadList:
                i.setDaemon(True)
                print("线程" + i.getName() + "：开始工作》》》》》")
                i.start()

            for j in [self.Data]:
                j.join()

            self.Book.save(self.dir + self.prefixName + self.suffixName)
            self.stopThread()
            print("07文件行级数据保存完毕")
        else:
            print("无可添加的行级数据")

    def write07Excel(self, data):
        while True:
            self.Loc.acquire()
            info = self.Data.get()
            if info:
                if len(data) == 0 or len(data) == 1:
                    self.LastColumn = 1
                else:
                    self.LastColumn = self.CurrentColumn

                for j in range(0, len(self.Field)):
                    key = self.Field[j]
                    self.Sheet.cell(self.LastRow, self.LastColumn, info.values[key])
                    self.LastColumn = self.LastColumn + 1

            self.LastRow = self.LastRow + 1

            self.Data.task_done()
            self.Loc.release()

    '''
        comm:  添加03版文件标题
        param: title[元组类型(字段1，字段2，字段3，……)]
    '''

    def addTitle03(self, row_id=1, column_id=1, title=()):

        if len(title) != 0:
            self.Field = title

        if self.Field is not None:
            print("添加03标题")
            self.Book = copy(xlrd.open_workbook(self.dir + self.prefixName + self.suffixName))

            self.Sheet = self.Book.get_sheet(0)
            self.LastRow = row_id - 1
            self.LastColumn = column_id - 1
            self.CurrentRow = row_id - 1
            self.CurrentColumn = column_id - 1
            for j in range(0, len(self.Field)):
                self.Sheet.write(self.LastRow, self.LastColumn, self.Field[j])
                self.LastColumn = self.LastColumn + 1
            self.Book.save(self.dir + self.prefixName + self.suffixName)
            self.LastRow = self.LastRow + 1

        else:
            print("无可添加的标题")

    '''
        comm:  03文件追加数据
        param: data[元组类型([DataRow],(字段1，字段2，字段3，……),文件路径,开始行号，开始列号)]
    '''

    def append03(self, *data):
        appendFileDir = self.dir + self.prefixName + self.suffixName

        if len(data) == 1:
            for j in data[0][0]:
                self.Data.put(j)

        if len(data) == 4:
            self.LastRow = data[0] - 1
            self.LastColumn = data[1] - 1
            self.CurrentRow = data[0] - 1
            self.CurrentColumn = data[1] - 1
            self.LocNum = data[2]
            self.Data = Queue()
            for j in data[3][0]:
                self.Data.put(j)

        if len(data) == 6:
            self.Data = Queue()
            for j in data[0]:
                self.Data.put(j)
            self.Field = data[1]
            appendFileDir = data[2]
            self.LastRow = data[3] - 1
            self.LastColumn = data[4] - 1
            self.CurrentRow = data[3] - 1
            self.CurrentColumn = data[4] - 1
            self.LocNum = data[5]

        if self.Field is not None and self.Data is not None:
            print("添加03文件的行级数据")
            self.Book = copy(xlrd.open_workbook(appendFileDir))
            self.Sheet = self.Book.get_sheet(0)

            for j in range(self.LocNum):
                excel = threading.Thread(target=self.write03Excel, args=(data,))
                self.ThreadList.append(excel)

            for i in self.ThreadList:
                i.setDaemon(True)
                print("线程" + i.getName() + "：开始工作》》》》》")
                i.start()

            for j in [self.Data]:
                j.join()

            self.Book.save(self.dir + self.prefixName + self.suffixName)
            self.stopThread()
            print("03文件行级数据保存完毕")
        else:
            print("无可添加的行级数据")

    def write03Excel(self, data):
        while True:
            self.Loc.acquire()
            info = self.Data.get()
            if info:
                if len(data) == 0 or len(data) == 1:
                    self.LastColumn = 0
                else:
                    self.LastColumn = self.CurrentColumn

                for j in range(0, len(self.Field)):
                    key = self.Field[j]
                    self.Sheet.write(self.LastRow, self.LastColumn, info.values[key])
                    self.LastColumn = self.LastColumn + 1

            self.LastRow = self.LastRow + 1

            self.Data.task_done()
            self.Loc.release()

    '''
        comm:  追加模板数据
        param: dir_info     模板文件路径
               row          追加的模板数据
               cloumn       
               row_id       开始行号
               column_id    开始列号
    '''

    def to_template(self, dir_info, row, cloumn, row_id, column_id,threadNum):
        if os.path.exists(dir_info):
            self.file_attr(dir_info)
            if self.suffixName == ".xls":
                print("开始追加03文件")
                self.append03(row, cloumn, dir_info, row_id, column_id,threadNum)
            if self.suffixName == ".xlsx":
                print("开始追加07文件")
                self.append07(row, cloumn, dir_info, row_id, column_id,threadNum)
        else:
            print("该模板文件不存在！")

    def stopThread(self):
        self.ThreadList.clear()
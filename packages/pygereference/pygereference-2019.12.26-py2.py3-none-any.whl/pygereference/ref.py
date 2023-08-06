#!/usr/bin/env python
# -*- coding: utf-8 -*-
# -----------------------------------------------------------------------------
# Le référentiel d'information de Guichet Entreprises est mis à disposition
# selon les termes de la licence Creative Commons Attribution - Pas de
# Modification 4.0 International.

# Pour accéder à une copie de cette licence, merci de vous rendre à l'adresse
# suivante :
# http://creativecommons.org/licenses/by-nd/4.0/
# ou envoyez un courrier à Creative Commons, 444 Castro Street, Suite 900,
# Mountain View, California, 94041, USA.
# -----------------------------------------------------------------------------

import logging
import sys
import os
import pickle
import re
import openpyxl

from pymdtools import common
from pymdtools import mdfile

__ref_filename__ = "reference"

__folder_name__ = ["%(category-filename-short-slug)s",
                   "%(domain-filename-short-slug)s",
                   "%(key)s-%(name-filename-short-slug)s"]

__filename_ref__ = "%(key)s-%(name-filename-short-slug)s"


# -----------------------------------------------------------------------------
# A decorator to cache the result of a function to the disk
# the function must be of the following type :
#    def the_function(filename)
#
# Could be usefull for an heavy initialisation from a file
# -----------------------------------------------------------------------------
@common.simple_decorator
def cache_on_disk(function_to_decorated):
    # helper
    def call_func_and_save_cache(filename, cache_filename):
        result = function_to_decorated(filename)
        file = open(cache_filename, "wb")
        logging.debug("Create cache %s", cache_filename)
        pickle.dump(result, file)
        file.close()
        return result

    # helper
    def read_cache(cache_filename):
        file = open(cache_filename, "rb")
        logging.debug("Read cache %s", cache_filename)
        result = pickle.load(file)
        file.close()
        return result

    # the return function
    def cached_init_from_file(filename, refresh=False):
        the_filename = common.set_correct_path(filename)
        cache_filename = the_filename + \
            "." + function_to_decorated.__name__ + \
            ".cache"

        if not os.path.isfile(cache_filename) or refresh:
            return call_func_and_save_cache(filename, cache_filename)

        # last modification time
        mtime_file = os.path.getmtime(the_filename)
        mtime_cache = os.path.getmtime(cache_filename)

        if mtime_file > mtime_cache:
            return call_func_and_save_cache(filename, cache_filename)
        return read_cache(cache_filename)

    return cached_init_from_file


# -----------------------------------------------------------------------------
# Get the reference excel file
#
# @param filename the information referential name
# @param extension extension of the filename
# @return the openxl workbook object
# -----------------------------------------------------------------------------
@common.static(__cache_result__=None)
def find_ref_filename(filename=__ref_filename__, extension=".xlsx"):
    logging.debug('Find the referential workbook')
    if find_ref_filename.__cache_result__ is None:
        find_ref_filename.__cache_result__ = {}

    key = filename + extension
    if key in find_ref_filename.__cache_result__:
        return find_ref_filename.__cache_result__[key]

    value = common.search_for_file(key, [__get_this_filename()],
                                   ["./"], nb_up_path=4)
    find_ref_filename.__cache_result__[key] = value

    return value


# -----------------------------------------------------------------------------
# Get the reference excel file
#
# @param filename the information referential name
# @param extension extension of the filename
# @return the openxl workbook object
# -----------------------------------------------------------------------------
def get_ref_workbook(filename=__ref_filename__, extension=".xlsx"):
    logging.debug('Load the referential workbook')

    the_real_filename = find_ref_filename(
        filename=filename, extension=extension)

    result = openpyxl.load_workbook(filename=the_real_filename,
                                    read_only=True)

    return result


# -----------------------------------------------------------------------------
# Check value from excel
# @param value the value
# @return the value corrected
# -----------------------------------------------------------------------------
def correct_xl_value(value):
    if value is None:
        return None

    result = value.strip()

    # result = result.replace(u'\u2019', "'")  # none breakable space
    # result = result.replace(u'\u2013', "-")  # long dash
    # result = result.replace(u'\u2026', "...")  # compressed ...

    return result

# -----------------------------------------------------------------------------
# Get the worksheet from the excel workbook
#
# @param worksheet_name name of the worksheet
# @param workbook the workbook
# @return the worksheet found
# -----------------------------------------------------------------------------
def get_ws_from_wb(worksheet_name, workbook):
    if worksheet_name not in workbook.sheetnames:
        logging.error('%s is not a worksheet of the referential',
                      worksheet_name)
        raise RuntimeError(
            '%s is not a worksheet of the referential' % worksheet_name)
    return workbook[worksheet_name]


# -----------------------------------------------------------------------------
# Get the worksheet position from position x,y
#
# @param row_num the position number start at 0
# @param col_num the position number start at 0
# @return the coordinate start at "A1"
# -----------------------------------------------------------------------------
def get_excel_coordinate(row_num, col_num):
    row_num += 1
    col_num += 1

    col_str = ''

    while col_num:
        remainder = col_num % 26

        if remainder == 0:
            remainder = 26

        # Convert the remainder to a character.
        col_letter = chr(ord('A') + remainder - 1)

        # Accumulate the column letters, right to left.
        col_str = col_letter + col_str

        # Get the next order of magnitude.
        col_num = int((col_num - 1) / 26)

    return col_str + str(row_num)


# -----------------------------------------------------------------------------
# find a column name
#
# @param column_name the name of the column to search for
# @param worksheet the workbook
# @param raw_number the line for search
# @return the column found or None
# -----------------------------------------------------------------------------
def find_col_from_ws(column_name, worksheet, raw_number=1):
    col_number_found = -1
    for i in range(0, worksheet.max_column):
        if worksheet[get_excel_coordinate(raw_number, i)].value == column_name:
            col_number_found = i

    if col_number_found == -1:
        return None

    return get_excel_coordinate(0, col_number_found)[:-1]


# -----------------------------------------------------------------------------
# Function to get the information referential
#
# @param name the name to transform
# @param lang the language
# @return the info referential
# -----------------------------------------------------------------------------
def get_name_trans(name, lang='fr'):
    result = {}
    expand = expand_abbreviations(name, lang=lang)
    if expand != name:
        result['expand'] = expand
        result['expand-ascii'] = common.str_to_ascii(expand)

    filename = common.get_valid_filename(name, replacement=" ")
    filename = re.subn(r"\s+", ' ', filename)[0]

    result['ascii'] = common.str_to_ascii(name)
    result['short'] = common.limit_str(name, 30, ' ')
    result['filename'] = filename
    result['filename-ascii'] = common.str_to_ascii(filename)
    result['filename-short'] = common.limit_str(filename, 30, ' ')
    result['filename-short-slug'] = common.slugify(result['filename-short'])

    return result


# -----------------------------------------------------------------------------
# Function to get the information referential
#
# @param referential the referential
# @return the info referential
# -----------------------------------------------------------------------------
def expand_ref(referential):
    result = referential

    for key1 in result:
        for key2 in result[key1]:
            words = list(result[key1][key2].keys())
            for word in words:
                expand_words = get_name_trans(result[key1][key2][word],
                                              lang=key2)
                for ext in expand_words:
                    result[key1][key2][word + '-' + ext] = expand_words[ext]

    for key1 in result:
        key_folder = ""
        for name in __folder_name__:
            key_folder = os.path.join(key_folder, name % result[key1]['fr'])
        result[key1]['path'] = key_folder
        result[key1]['filename'] = __filename_ref__ % result[key1]['fr']
        result[key1]['fr']['path'] = key_folder
        result[key1]['fr']['filename_gen'] = \
            __filename_ref__ % result[key1]['fr']
        result[key1]['fr']['filename'] = \
            __filename_ref__ % result[key1]['fr'] + ".fr.md"

        if 'en' not in result[key1]:
            result[key1]['en'] = {}
        result[key1]['en']['path'] = key_folder
        result[key1]['en']['filename'] = \
            __filename_ref__ % result[key1]['fr'] + ".en.md"

        if 'src' not in result[key1]:
            result[key1]['src'] = {}
        result[key1]['src']['path'] = key_folder
        result[key1]['src']['filename'] = \
            __filename_ref__ % result[key1]['fr'] + ".sources.md"

    return result


# -----------------------------------------------------------------------------
# Function to get the information referential
#
# @param key the key to find the info data (if key is None, return the dict)
# @return the info referential
# -----------------------------------------------------------------------------
@common.static(__info_ref_dict__=None)
def get_info_ref(key=None, lang="fr"):
    if get_info_ref.__info_ref_dict__ is None:
        filename = find_ref_filename(__ref_filename__)
        ref = build_dict_from_file(filename)['Ref']
        get_info_ref.__info_ref_dict__ = expand_ref(ref)

    if key is not None:
        key = key.upper()
        if lang is not None:
            return get_info_ref.__info_ref_dict__[key][lang]
        return get_info_ref.__info_ref_dict__[key]

    return get_info_ref.__info_ref_dict__


# -----------------------------------------------------------------------------
# Function to get the information referential
#
# @param key the key to find the info data (if key is None, return the dict)
# @return the info referential
# -----------------------------------------------------------------------------
@common.static(__info_ref_dict__=None)
def expand_abbreviations(key=None, lang="fr"):
    if expand_abbreviations.__info_ref_dict__ is None:
        filename = find_ref_filename(__ref_filename__)
        expand_abbreviations.__info_ref_dict__ = \
            build_dict_from_file(filename)['Abr']

    if key is not None:
        key_upper = key.upper()
        if key_upper not in expand_abbreviations.__info_ref_dict__:
            return key
        if lang not in expand_abbreviations.__info_ref_dict__[key_upper]:
            return key
        return expand_abbreviations.__info_ref_dict__[
            key_upper][lang]['expand']

    return expand_abbreviations.__info_ref_dict__


# -----------------------------------------------------------------------------
# Get info data
#
# @return the key list
# -----------------------------------------------------------------------------
def get_info_ref_keys():
    return list(get_info_ref().keys())


# -----------------------------------------------------------------------------
# Get info data
#
# @return the key list
# -----------------------------------------------------------------------------
def get_info_ref_filename(base_path, key, lang='fr'):
    temp = get_info_ref(key, lang=lang)
    return os.path.join(base_path, temp['path'], temp['filename'])

# -----------------------------------------------------------------------------
# List all the domain
# -----------------------------------------------------------------------------
def get_domain_list(lang='fr'):
    keys = get_info_ref_keys()
    result = {}

    for k in keys:
        info = get_info_ref(k, lang=lang)
        category = info['category']
        domain = info['domain']
        if category not in result:
            result[category] = []
        if domain not in result[category]:
            result[category].append(domain)

    return result


# -----------------------------------------------------------------------------
# Initialize the keys of the referential information
#
# @param filename the filename
# @return the dict build
# -----------------------------------------------------------------------------
@cache_on_disk
def build_dict_from_file(filename):
    the_real_filename = common.check_is_file_and_correct_path(filename)
    workbook = openpyxl.load_workbook(filename=the_real_filename,
                                      read_only=True)
    return build_dict_from_xl_wb(workbook)


# -----------------------------------------------------------------------------
# Build dict from excelfile
#
# @param workbook the workbook
# @param key1 the first key
# @param key2 the second key
# @return the dict build
# -----------------------------------------------------------------------------
def build_dict_from_xl_wb(workbook, key1="Key", key2="Lang"):
    result = {}
    ws_list = workbook.sheetnames

    for worksheet_name in ws_list:
        worksheet = get_ws_from_wb(worksheet_name, workbook)
        result[worksheet_name] = build_dict_from_xl_ws(
            worksheet, key1=key1, key2=key2)

    return result


# -----------------------------------------------------------------------------
# Build dict from excelfile
#
# @param worksheet the worksheet
# @param key1 the first key
# @param key2 the second key
# @return the dict build
# -----------------------------------------------------------------------------
def build_dict_from_xl_ws(worksheet, key1="Key", key2="Lang"):
    result = {}
    raw_title = 0
    col_list = {}
    key1 = key1.lower()
    key2 = key2.lower()

    for i in range(0, worksheet.max_column):
        value = worksheet[get_excel_coordinate(raw_title, i)].value
        value = correct_xl_value(value)

        if (value is not None) and (len(value) > 1):
            col_list[value.lower()] = i

    if (key1 not in col_list) or (key2 not in col_list):
        logging.error('Can not find the key %s %s in the worksheet',
                      key1, key2)
        raise RuntimeError('Can not find the key %s %s in the worksheet' %
                           (key1, key2))

    logging.info('Read the key until ligne %s', worksheet.max_row)
    for i in range(raw_title + 1, worksheet.max_row):
        value = {}
        for col_name in col_list:
            xl_coord = get_excel_coordinate(i, col_list[col_name])
            value[col_name] = correct_xl_value(worksheet[xl_coord].value)

        if (value[key1] is None) or (value[key2] is None):
            break

        logging.info('Read line %d the key=%s', i, value[key1])

        value[key1] = re.subn(r"\s", "", value[key1])[0].upper()
        value[key2] = re.subn(r"\s", "", value[key2])[0].lower()

        if value[key1] not in result:
            result[value[key1]] = {}
        if value[key2] not in result[value[key1]]:
            result[value[key1]][value[key2]] = {}

        for key in value:
            result[value[key1]][value[key2]][key] = value[key]

    return result


# -----------------------------------------------------------------------------
# Find the filename of this file (depend on the frozen or not)
# This function return the filename of this script.
# The function is complex for the frozen system
#
# @return the folder of THIS script.
# -----------------------------------------------------------------------------
def __get_this_folder():
    return os.path.split(os.path.abspath(os.path.realpath(
        __get_this_filename())))[0]

# -----------------------------------------------------------------------------
# Find the filename of this file (depend on the frozen or not)
# This function return the filename of this script.
# The function is complex for the frozen system
#
# @return the filename of THIS script.
# -----------------------------------------------------------------------------
def __get_this_filename():
    result = ""

    if getattr(sys, 'frozen', False):
        # frozen
        result = sys.executable
    else:
        # unfrozen
        result = __file__

    return result


# -----------------------------------------------------------------------------
__REF_FOLDER__ = {}
# -----------------------------------------------------------------------------
def get_ref_folder():
    if 'path' not in __REF_FOLDER__:
        # make a guess
        ref_folder = os.path.join(__get_this_folder(), "..", "content/")
        __REF_FOLDER__['path'] = ref_folder

    return __REF_FOLDER__['path']

# -----------------------------------------------------------------------------
def set_ref_folder(the_new_path):
    __REF_FOLDER__['path'] = common.check_folder(the_new_path)
    return __REF_FOLDER__['path']

# -----------------------------------------------------------------------------
@cache_on_disk
def create_ref_from_files(folder):
    result = {}

    def collect(filename):
        logging.debug("collect %s", filename)
        md = mdfile.MarkdownContent(filename)
        if 'key' not in md.keys() or 'lang' not in md.keys():
            return
        key = md['key'].upper()
        lang = md['lang'].lower()
        if key not in result:
            result[key] = {}
        if lang in result[key]:
            raise Exception("Two file for the key %s lang %s" % (key, lang))
        result[key][lang] = {}
        for name, value in md.items():
            result[key][lang][name] = value
        result[key][lang]['filename'] = md.full_filename
        logging.info("collect key=%s lang=%s %s ", key, lang, filename)

    common.apply_function_in_folder(folder, collect, filename_ext=".md")

    return result

# -----------------------------------------------------------------------------
# Function to get the information referential
#
# @param key the key to find the info data (if key is None, return the dict)
# @return the info referential
# -----------------------------------------------------------------------------
@common.static(__info_ref_dict__=None)
def get_ref(key=None, lang=None, refresh=False):
    if get_ref.__info_ref_dict__ is None:
        if get_ref_folder() is None:
            raise Exception("You must set the folder for reference")
        get_ref.__info_ref_dict__ = \
            create_ref_from_files(get_ref_folder(), refresh=refresh)

    if key is not None:
        key = key.upper()
        if lang is not None:
            return get_ref.__info_ref_dict__[key][lang]
        return get_ref.__info_ref_dict__[key]

    return get_ref.__info_ref_dict__

# -----------------------------------------------------------------------------
# Function to get the information referential
#
# @param key the key to find the info data (if key is None, return the dict)
# @return the info referential
# -----------------------------------------------------------------------------
def get_ref_langs():
    refs = get_ref()
    result = []
    for key in refs:
        for lang in refs[key]:
            if lang not in result:
                result.append(lang)
    return result

# -----------------------------------------------------------------------------
# Function to get the information referential
#
# @param key the key to find the info data (if key is None, return the dict)
# @return the info referential
# -----------------------------------------------------------------------------
def get_ref_organised(lang, keys_order):
    refs = get_ref()
    temp = {}
    for key in refs:
        if lang not in refs[key]:
            continue
        temp[key] = {
            'filename': refs[key][lang]['filename'],
            'title': refs[key][lang]['title'],
            'key': key,
        }
        for level in keys_order:
            temp[key][level] = refs[key][lang][level]

    result = {}
    for key in temp:
        inter = result
        for level in keys_order:
            if temp[key][level] not in inter:
                inter[temp[key][level]] = {}
            inter = inter[temp[key][level]]
        inter[key] = {
            'filename': temp[key]['filename'],
            'title': temp[key]['title'],
            'key': key,
        }

    return result

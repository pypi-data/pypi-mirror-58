#!/usr/bin/env python
# -*- coding: utf-8  -*-

import os.path
import re

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

from datalet.data import *
from datalet.storage.bin_file_storage import BinFileStorage
from datalet.storage.exceptions import *


class ExcelXlsxStorage(BinFileStorage):
	"""
	Excel 2007 format Storage
	"""


	def __init__(self, location = None, sheet_index = -1, sheet_name = None):
		super().__init__(location)
		self.sheet_index = sheet_index
		self.sheet_name = sheet_name


	def __get_worksheet(self, workbook_toread):
		ws = None
		if not self.sheet_name is None:
			ws = workbook_toread.get_sheet_by_name(self.sheet_name)
		else:
			if not self.sheet_index is None:
				ws = workbook_toread.get_sheet_by_name(workbook_toread.get_sheet_names()[self.sheet_index])
			else:
				raise ArgumentsAbsenceError("The targetSheet arguments must be specified one:  sheet_index=%s, sheet_name=%s" % \
					(self.sheet_index, self.sheet_name))
		return ws


	def create(self, force = False):
		if self.exists():
			if force == False:
				raise StorageExistsError(self.location)
			else:
				self.remove(force = True)

		wb = Workbook()
		new_sheet_name = self.sheet_name if self.sheet_name is not None else ("SHEET_" + str(self.sheet_index))
		wb.create_sheet(title = new_sheet_name)
		wb.save(self.location)


	def clear(self, force = False):
		wb = load_workbook(filename = self.location)
		ws = self.__get_worksheet(wb)
		wb.remove(ws)
		new_sheet_name = self.sheet_name if self.sheet_name is not None else ("SHEET_" + str(self.sheet_index))
		wb.create_sheet(title = new_sheet_name)
		wb.save(self.location)


	def read(self, limit =  -1, data_only = True, read_only = True, encoding='utf-8'):
		dt = DataTable()
		'''
		read_only: set True when read large file.
		data_only: set True when need calculate the formatular.
		'''
		wb = load_workbook(filename = self.location, data_only = data_only, read_only = read_only)
		ws = self.__get_worksheet(wb)

		for row in ws.rows:
			dr = DataRow()
			for cell in row:
				dr.append(cell.value)
			dt.append(dr)

		# get cell data
		#for rowNum in range(1, ws.max_row + 1):
		#	if limit > -1 and rowNum > limit:
		#		break

		#	rowdata = []
		#	for colNum in range(1, ws.max_column + 1):
		#		cellval = ws.cell(row = rowNum, column= colNum).value
		#		rowdata.append(cellval)
		#	tabledata.append(rowdata)

		return dt

	def write(self, data, force = True, overwrite = False, include_header = True, encoding = 'utf-8', write_only = True):
		'''
		use write_only mode to write large data to file.
		'''
		ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
		# if overwrite == True:
			# self.clear(force = True)
		if write_only == True:
			# data = [] if data is None else data
			# if self.exists():
				# old_data = self.read()
				# old_data.extend(data)
				# data = old_data
			#self.remove(force = True)
			wb = Workbook(write_only = True)
			ws = wb.create_sheet(title = self.sheet_name)
			if include_header == True:
				ws.append(data.column_names)
			for row in data:
				new_row = []
				for data in row:
					new_data = str(data) if data is not None else ''
					if ILLEGAL_CHARACTERS_RE.match(new_data) is not None:
						new_data = "*"
					new_row.append(new_data)
				try:
					ws.append(list(map(str, new_row)))
				except Exception as e:
					print(row)
					print(e)
					raise e
			wb.save(self.location)
		else:
			wb = load_workbook(filename = self.location)
			ws = self.__get_worksheet(wb)
			datarow_start_index = 1
			if include_header == True:
				for col_header_index in range(0, len(data.column_names)):
					cell_header_data = data.column_names[col_header_index]
					ws.cell(row = datarow_start_index, column = col_header_index + 1).value = cell_header_data
				datarow_start_index += 1
			for rowIndex in range(0, len(data)):
				rowdata = data[rowIndex]
				for colIndex in range(0, len(rowdata)):
					celldata = rowdata[colIndex]
					ws.cell(row = rowIndex + datarow_start_index, column = colIndex + 1).value = celldata
			wb.save(self.location)
